"""
master-pull.py
==============================================================================
Rebuilds ``DataInflShortLongConsUpdated.xlsx`` -> ``DataInflShortLongConsUpdated_2025.xlsx``
and ``DataInflShortLongUpdated_2025.xlsx`` (same data, without ``rconpc_*`` columns).

Rule (per user request):
  * For every column, keep the existing workbook values for years <= 2016.
  * For years 2017..2025, REPLACE with freshly fetched data, using the exact
    source/method demonstrated in each country's notebook under ``data/<country>/``.

Series conventions (matching the original workbook):
  * cpi_*    : source index rebased to 1990 = 100; placed directly for 2017-2019.
               If the source level at 2020 differs from the existing workbook by
               more than CHAIN_REANCHOR_PCT (0.5%), pin 2020 to the existing value
               and chain 2021+ from that 2020 anchor using source percent changes.
  * rconpc_* : World Bank NE.CON.PRVT.PC.KD rebased to 2010 = 100; same 2020
               re-chain rule as cpi_*.
  * stir_* / ltir_* / baa_usa : annual rate values placed directly.

Columns with no defined fetch source are left untouched (existing values kept):
  (none — all rate columns are refreshed from 2017+).

Robustness: each series fetch is wrapped in try/except.  A fetch that returns at
least one 2017-2025 value is treated as a success and overwrites all of
2017-2025 (blank where the source has no value).  A fetch that fails outright or
returns nothing leaves the existing values in place (and is reported).
"""

import io
import json
import ssl
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import warnings

import pandas as pd

from api_keys import BDF_KEY, FRED_KEY  

warnings.simplefilter("ignore")

# --------------------------------------------------------------------------- #
# Paths / constants
# --------------------------------------------------------------------------- #
import os

HERE = os.path.dirname(os.path.abspath(__file__))
INDATA = os.path.abspath(os.path.join(HERE, "..", "..", "indata"))
SRC_XLSX = os.path.join(INDATA, "DataInflShortLongConsUpdated.xlsx")
OUT_XLSX = os.path.join(INDATA, "DataInflShortLongConsUpdated_2025.xlsx")
OUT_XLSX_NO_RCONPC = os.path.join(INDATA, "DataInflShortLongUpdated_2025.xlsx")
RBA_CSV = os.path.join(INDATA, "raw", "f1.1-data.csv")

ANCHOR = 2020          # last year taken from the existing workbook
CPI_BASE_YEAR = 1990   # cpi source index with this year = 100
RCONPC_BASE_YEAR = 2010  # rconpc source index with this year = 100
CHAIN_REANCHOR_YEAR = 2020   # cpi/rconpc: optional re-chain anchor year
CHAIN_REANCHOR_PCT = 0.5     # re-chain 2021+ from existing 2020 if source 2020 differs by more than this %
FETCH_START = 2021     # first year to (re)fetch
FETCH_END = 2025       # last (complete) year to fetch
SR_OECD_FROM = 2025    # stir_usa / stir_uk: MeasuringWorth before this, OECD IR3TIB from here


UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

# ISO-3 codes keyed by workbook 2-letter suffix
ISO3 = {
    "usa": "USA", "deu": "DEU", "uk": "GBR", "fr": "FRA", "ca": "CAN",
    "it": "ITA", "jp": "JPN", "au": "AUS", "be": "BEL", "fi": "FIN",
    "ie": "IRL", "nl": "NLD", "no": "NOR", "ch": "CHE", "se": "SWE",
    "es": "ESP", "pt": "PRT", "dk": "DNK",
}

_SSL = ssl.create_default_context()
_SSL.check_hostname = False
_SSL.verify_mode = ssl.CERT_NONE


def _get(url, headers=None, timeout=120, data=None, retries=6):
    """HTTP GET/POST with retry+backoff on 429 (rate-limit) and 5xx errors."""
    h = {"User-Agent": UA, "Accept": "*/*"}
    if headers:
        h.update(headers)
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=h, data=data)
            with urllib.request.urlopen(req, timeout=timeout, context=_SSL) as resp:
                return resp.read()
        except urllib.error.HTTPError as exc:
            last = exc
            if exc.code == 429 or exc.code >= 500:
                wait = exc.headers.get("Retry-After") if exc.headers else None
                delay = float(wait) if (wait and str(wait).isdigit()) else 3 * (attempt + 1)
                time.sleep(min(delay, 30))
                continue
            raise
        except Exception as exc:  # noqa (timeouts, conn resets)
            last = exc
            time.sleep(3 * (attempt + 1))
    raise last


def _annual_mean_from_monthly(months, values):
    m = pd.DataFrame({"month": months, "v": pd.to_numeric(values, errors="coerce")}).dropna()
    m["year"] = m["month"].dt.year
    return m.groupby("year")["v"].mean()


def _rebase_index(series, base_year):
    """Rebase a level/index series so ``base_year`` = 100."""
    s = series.dropna().sort_index()
    if base_year not in s.index:
        return s
    base = float(s[base_year])
    if abs(base) < 1e-9:
        return s
    return s * (100.0 / base)


# --------------------------------------------------------------------------- #
# CPI level fetchers  ->  Series(index=year, value=annual level, 1990 = 100)
# --------------------------------------------------------------------------- #
def cpi_weo(iso3):
    """IMF WEO PCPIPCH: chain annual inflation from 1990 = 100."""
    url = (f"https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.RES/WEO/+/"
           f"{iso3}.PCPIPCH.A?startPeriod={CPI_BASE_YEAR}&format=csv")
    rows = list(csv_reader(_get(url, {"Accept": "text/csv"})))
    infl = {int(r["TIME_PERIOD"]): float(r["OBS_VALUE"])
            for r in rows if r.get("OBS_VALUE") not in (None, "")}
    level = {CPI_BASE_YEAR: 100.0}
    for y in range(CPI_BASE_YEAR + 1, FETCH_END + 1):
        if y in infl:
            level[y] = level[y - 1] * (1 + infl[y] / 100.0)
    return pd.Series(level)


def cpi_imf_ifs_fr():
    """IMF IFS CPI, All items (monthly -> annual average), 1990 = 100."""
    url = ("https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.STA/CPI/+/"
           f"FRA.CPI._T.IX.M?startPeriod={CPI_BASE_YEAR}-01&format=csv")
    rows = list(csv_reader(_get(url, {"Accept": "text/csv"})))
    months = pd.to_datetime([r["TIME_PERIOD"] for r in rows], format="%Y-M%m")
    annual = _annual_mean_from_monthly(months, [r["OBS_VALUE"] for r in rows])
    return _rebase_index(annual, CPI_BASE_YEAR)


def cpi_bls_usa():
    """BLS R-CPI-U-RS annual average (AVG), 1990 = 100."""
    raw = _get("https://www.bls.gov/cpi/research-series/r-cpi-u-rs-allitems.xlsx")
    df = pd.read_excel(io.BytesIO(raw), skiprows=5)
    df.columns = [str(c).strip().upper() for c in df.columns]
    df = df[["YEAR", "AVG"]].copy()
    df["YEAR"] = pd.to_numeric(df["YEAR"], errors="coerce")
    df["AVG"] = pd.to_numeric(df["AVG"], errors="coerce")
    df = df.dropna()
    s = pd.Series(df["AVG"].values, index=df["YEAR"].astype(int).values)
    return _rebase_index(s, CPI_BASE_YEAR)


def cpi_ons_uk():
    """ONS D7BT CPI index (annual), 1990 = 100."""
    raw = _get("https://www.ons.gov.uk/generator?format=csv&uri=/economy/"
               "inflationandpriceindices/timeseries/d7bt/mm23").decode()
    out = {}
    for line in raw.splitlines():
        parts = [p.strip('"') for p in line.split(",")]
        if len(parts) >= 2 and parts[0].isdigit() and len(parts[0]) == 4:
            out[int(parts[0])] = float(parts[1])
    return _rebase_index(pd.Series(out), CPI_BASE_YEAR)


def cpi_cso_ie():
    """CSO Ireland CPA04 annual index, 1990 = 100."""
    raw = _get("https://ws.cso.ie/public/api.restful/PxStat.Data.Cube_API."
               "ReadDataset/CPA04/CSV/1.0/en").decode("utf-8-sig", "replace")
    df = pd.read_csv(io.StringIO(raw))
    df = df[(df["C02071V02502"] == 12) & df["VALUE"].notna()]
    s = pd.Series(df["VALUE"].astype(float).values, index=df["Year"].astype(int).values)
    return _rebase_index(s, CPI_BASE_YEAR)


def cpi_scb_se():
    """Statistics Sweden KPILevindexM monthly -> annual average, 1990 = 100."""
    payload = json.dumps({
        "query": [
            {"code": "ContentsCode", "selection": {"filter": "item", "values": ["PR0101A4"]}},
            {"code": "Tid", "selection": {"filter": "all", "values": ["*"]}},
        ],
        "response": {"format": "json-stat2"},
    }).encode()
    res = json.loads(_get("https://api.scb.se/OV0104/v1/doris/en/ssd/START/PR/PR0101/PR0101A/KPILevindexM",
                          {"Content-Type": "application/json"}, data=payload))
    labels = list(res["dimension"]["Tid"]["category"]["label"].values())
    months = pd.to_datetime(labels, format="%YM%m")
    monthly = pd.DataFrame({"month": months, "cpi": res["value"]}).dropna(subset=["cpi"])
    annual = monthly.assign(year=monthly["month"].dt.year).groupby("year")["cpi"].mean()
    base = monthly.loc[monthly["month"].dt.year == CPI_BASE_YEAR, "cpi"].mean()
    return annual * (100.0 / base)


def cpi_norway():
    """Norges Bank HMS (<=2017) + SSB 03014 (>=2018), 1990 = 100."""
    nb_raw = _get("https://www.norges-bank.no/globalassets/upload/hms/data/cpi.xlsx")
    nb = pd.ExcelFile(io.BytesIO(nb_raw)).parse("p1_c3_table_14", header=None).iloc[20:].copy()
    nb.columns = ["year", "cpi", "wpi"]
    nb["year"] = pd.to_numeric(nb["year"], errors="coerce")
    nb["cpi"] = pd.to_numeric(nb["cpi"], errors="coerce")
    nb = nb.dropna(subset=["year", "cpi"])
    nb = nb[nb["year"] <= 2017]
    s = {int(y): float(c) for y, c in zip(nb["year"], nb["cpi"])}

    payload = json.dumps({
        "query": [
            {"code": "Konsumgrp", "selection": {"filter": "item", "values": ["TOTAL"]}},
            {"code": "ContentsCode", "selection": {"filter": "item", "values": ["KpiAar"]}},
            {"code": "Tid", "selection": {"filter": "all", "values": ["*"]}},
        ],
        "response": {"format": "json-stat2"},
    }).encode()
    res = json.loads(_get("https://data.ssb.no/api/v0/en/table/03014",
                          {"Content-Type": "application/json"}, data=payload))
    labels = list(res["dimension"]["Tid"]["category"]["label"].values())
    for y, v in zip(labels, res["value"]):
        if int(y) >= 2018 and v is not None:
            s[int(y)] = float(v)
    return _rebase_index(pd.Series(s).sort_index(), CPI_BASE_YEAR)


import csv as _csv


def csv_reader(raw_bytes):
    return _csv.DictReader(io.StringIO(raw_bytes.decode()))


# --------------------------------------------------------------------------- #
# Rate fetchers  ->  Series(index=year, value=annual rate %)
# --------------------------------------------------------------------------- #
def imf_mfs(iso3, indicator):
    """IMF IFS MFS_IR monthly rate -> calendar-year average."""
    url = (f"https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.STA/MFS_IR/+/"
           f"{iso3}.{indicator}.M?startPeriod={FETCH_START}-01&format=csv")
    rows = [r for r in csv_reader(_get(url, {"Accept": "text/csv"})) if r.get("OBS_VALUE")]
    if not rows:
        return pd.Series(dtype=float)
    months = pd.to_datetime([r["TIME_PERIOD"] for r in rows], format="%Y-M%m")
    return _annual_mean_from_monthly(months, [r["OBS_VALUE"] for r in rows])


def imf_bonds_then_oecd(iso3):
    """Long rate: IMF IFS Government Bonds (user's chosen source) where available,
    filled with OECD 10-yr yields for any 2017-2025 years IMF does not cover.
    (Switzerland/USA have sparse IMF coverage; OECD keeps the column complete.)"""
    try:
        imf = imf_mfs(iso3, "S13BOND_RT_PT_A_PT")
    except Exception:  # noqa
        imf = pd.Series(dtype=float)
    try:
        oecd = oecd_finmark(iso3, "IRLT")
    except Exception:  # noqa
        oecd = pd.Series(dtype=float)
    years = sorted(set(imf.index) | set(oecd.index))
    out = {}
    for y in years:
        if y in imf.index and pd.notna(imf[y]):
            out[y] = float(imf[y])
        elif y in oecd.index and pd.notna(oecd[y]):
            out[y] = float(oecd[y])
    return pd.Series(out).sort_index()


def imf_mm_then_oecd(iso3):
    """Short rate: IMF IFS money-market rate (user's chosen source) where available,
    filled with the OECD 3-month interbank rate for any 2017-2025 years IMF does
    not cover.  (Spain's IMF money-market series is discontinued after 2021.)"""
    try:
        imf = imf_mfs(iso3, "MMRT_RT_PT_A_PT")
    except Exception:  # noqa
        imf = pd.Series(dtype=float)
    try:
        oecd = oecd_short(iso3)
    except Exception:  # noqa
        oecd = pd.Series(dtype=float)
    years = sorted(set(imf.index) | set(oecd.index))
    out = {}
    for y in years:
        if y in imf.index and pd.notna(imf[y]):
            out[y] = float(imf[y])
        elif y in oecd.index and pd.notna(oecd[y]):
            out[y] = float(oecd[y])
    return pd.Series(out).sort_index()


# OECD is aggressively rate-limited, so every measure is fetched exactly ONCE for
# ALL countries at a time.  Results are cached in-memory (one network attempt per
# run) AND on disk (so a successful batch is reused by later runs without
# re-hitting OECD, which otherwise puts this IP in a 429 penalty box).
_OECD_CACHE = {}
_CACHE_DIR = os.path.join(HERE, "_cache")


def _cache_load(key):
    import pickle
    path = os.path.join(_CACHE_DIR, key + ".pkl")
    if os.path.exists(path):
        try:
            with open(path, "rb") as f:
                return pickle.load(f)
        except Exception:  # noqa
            return None
    return None


def _cache_save(key, obj):
    import pickle
    os.makedirs(_CACHE_DIR, exist_ok=True)
    with open(os.path.join(_CACHE_DIR, key + ".pkl"), "wb") as f:
        pickle.dump(obj, f)


def _oecd_finmark_batch(measure):
    """DF_FINMARK for ALL workbook countries in a single request -> {iso3: Series}.
    One network attempt per run; successful non-empty results persist to disk."""
    if measure in _OECD_CACHE:
        return _OECD_CACHE[measure]
    disk = _cache_load(f"oecd_finmark_{measure}")
    if disk:
        _OECD_CACHE[measure] = disk
        print(f"    (OECD {measure}: loaded {len(disk)} areas from disk cache)")
        return disk
    out = {}
    try:
        areas = "+".join(sorted(set(ISO3.values())))
        url = ("https://sdmx.oecd.org/public/rest/data/OECD.SDD.STES,DSD_STES@DF_FINMARK,4.0/"
               f"{areas}.M.{measure}.PA.....?startPeriod={FETCH_START}-01"
               "&dimensionAtObservation=AllDimensions&format=csvfilewithlabels")
        raw = _get(url, {"Accept": "application/vnd.sdmx.data+csv; charset=utf-8"})
        df = pd.read_csv(io.StringIO(raw.decode("utf-8")))
        if "OBS_VALUE" in df and "REF_AREA" in df and not df.empty:
            df = df.dropna(subset=["OBS_VALUE"])
            df["_m"] = pd.to_datetime(df["TIME_PERIOD"], errors="coerce")
            for area, g in df.groupby("REF_AREA"):
                out[area] = _annual_mean_from_monthly(g["_m"], g["OBS_VALUE"])
    except Exception as exc:  # noqa
        print(f"    (OECD DF_FINMARK {measure} batch failed: {exc})")
    _OECD_CACHE[measure] = out
    if out:
        _cache_save(f"oecd_finmark_{measure}", out)
    return out


def _mei_fin_all():
    """Legacy OECD MEI_FIN IR3TIB for all areas in one request; {iso3: Series}."""
    if "MEI_FIN" in _OECD_CACHE:
        return _OECD_CACHE["MEI_FIN"]
    disk = _cache_load("oecd_mei_fin")
    if disk:
        _OECD_CACHE["MEI_FIN"] = disk
        return disk
    out = {}
    try:
        raw = _get("https://stats.oecd.org/SDMX-JSON/data/MEI_FIN/IR3TIB..M/OECD?contentType=csv",
                   retries=3)
        df = pd.read_csv(io.StringIO(raw.decode("utf-8-sig", "replace")))
        df = df[(df["MEASURE"] == "IR3TIB")
                & (df["TIME_PERIOD"].astype(str).str.match(r"^\d{4}-\d{2}$"))]
        for area, g in df.groupby("REF_AREA"):
            months = pd.to_datetime(g["TIME_PERIOD"], format="%Y-%m")
            out[area] = _annual_mean_from_monthly(months, g["OBS_VALUE"])
    except Exception as exc:  # noqa
        print(f"    (MEI_FIN batch unavailable: {exc}; using DF_FINMARK IR3TIB)")
    _OECD_CACHE["MEI_FIN"] = out
    if out:
        _cache_save("oecd_mei_fin", out)
    return out


def oecd_finmark(iso3, measure):
    """OECD long/short rate for one country (from the cached all-country batch)."""
    return _oecd_finmark_batch(measure).get(iso3, pd.Series(dtype=float))


def oecd_short(iso3):
    """OECD short-term 3-month interbank rate: legacy MEI_FIN if it has the country,
    else modern DF_FINMARK IR3TIB.  Both are single cached batch fetches."""
    mei = _mei_fin_all().get(iso3, pd.Series(dtype=float))
    if len(mei) and mei[mei.index >= FETCH_START].notna().any():
        return mei
    return oecd_finmark(iso3, "IR3TIB")


def sr_measuringworth_usa():
    """MeasuringWorth US through 2024; OECD IR3TIB from 2025 (MeasuringWorth lags)."""
    return _sr_measuringworth_then_oecd("USA", table_idx=1, params={"data7": "on"})


def sr_measuringworth_uk():
    """MeasuringWorth UK through 2024; OECD IR3TIB from 2025 (MeasuringWorth lags)."""
    return _sr_measuringworth_then_oecd("GBR", table_idx=0, params={"data5": "on"})


def _sr_measuringworth_then_oecd(iso3, table_idx, params):
    """MeasuringWorth surplus-funds short rate through SR_OECD_FROM-1; OECD from SR_OECD_FROM."""
    mw = _measuringworth_stir(table_idx, params)
    mw = mw[mw.index < SR_OECD_FROM]
    ext = oecd_short(iso3)
    ext = ext[ext.index >= SR_OECD_FROM]
    return pd.concat([mw, ext]).sort_index()


def _measuringworth_stir(table_idx, params):
    import requests
    base_params = {"year_source": 1870, "year_result": 2030}
    base_params.update(params)
    r = requests.get("https://www.measuringworth.com/datasets/interestrates/result.php",
                     params=base_params, timeout=90)
    r.raise_for_status()
    tables = pd.read_html(io.StringIO(r.text))
    raw = tables[table_idx]
    raw.columns = raw.iloc[0]
    raw = raw.drop(index=0)
    raw = raw.rename(columns={raw.columns[0]: "year"})
    raw["year"] = pd.to_numeric(raw["year"], errors="coerce")
    raw = raw.dropna(subset=["year"])
    col = [c for c in raw.columns if c != "year"][0]
    return pd.Series(pd.to_numeric(raw[col], errors="coerce").values,
                     index=raw["year"].astype(int).values)


def sr_germany():
    """germany_sr.ipynb: EONIA (FRED EONIARATE) then EUR-STR + 0.085 pp, cal-yr avg."""
    frames = {}
    for sid, (start, end, adj) in {
        "EONIARATE": ("2012-01-01", "2021-12-31", 0.0),
        "ECBESTRVOLWGTTRMDMNRT": ("2022-01-01", "2099-12-31", 0.085),
    }.items():
        url = ("https://api.stlouisfed.org/fred/series/observations"
               f"?series_id={sid}&api_key={FRED_KEY}&file_type=json"
               f"&observation_start={start}&observation_end={end}")
        obs = json.loads(_get(url, {"Accept": "application/json"}))["observations"]
        d = pd.DataFrame({"date": [o["date"] for o in obs],
                          "v": pd.to_numeric([o["value"] for o in obs], errors="coerce")})
        d["date"] = pd.to_datetime(d["date"])
        d = d.dropna()
        ann = d.assign(year=d["date"].dt.year).groupby("year")["v"].mean() + adj
        for y, v in ann.items():
            frames[int(y)] = v
    return pd.Series(frames).sort_index()


def sr_france():
    url = ("https://webstat.banque-france.fr/api/explore/v2.1/catalog/datasets/"
           "observations/exports/json?" + urllib.parse.urlencode({
               "order_by": "time_period_start",
               "refine": 'series_key:"ECOFI.INR.FR.FITB_PA._Z.D"',
               "where": f'time_period_start >= "{FETCH_START}-01-01"',
           }))
    rows = json.loads(_get(url, {"Authorization": f"Apikey {BDF_KEY}",
                                 "Accept": "application/json"}))
    d = pd.DataFrame({"date": pd.to_datetime([r["time_period_start"] for r in rows]),
                      "v": pd.to_numeric([r["obs_value"] for r in rows], errors="coerce")}).dropna()
    return d.assign(year=d["date"].dt.year).groupby("year")["v"].mean()


def _boc_valet(series, start, end):
    url = (f"https://www.bankofcanada.ca/valet/observations/{series}/json"
           f"?start_date={start}-01-01&end_date={end}-12-31")
    payload = json.loads(_get(url, {"Accept": "application/json"}))
    rows = []
    for o in payload["observations"]:
        pt = o.get(series) or {}
        v = pt.get("v")
        if v not in (None, ""):
            rows.append((o["d"], float(v)))
    if not rows:
        return pd.DataFrame(columns=["date", "v"])
    df = pd.DataFrame(rows, columns=["date", "v"])
    df["date"] = pd.to_datetime(df["date"])
    return df


def sr_canada():
    df = _boc_valet("V80691303", FETCH_START, FETCH_END)          # 3-mo T-bill auction (weekly)
    monthly = df.sort_values("date").set_index("date")["v"].resample("ME").last().reset_index()
    return _annual_mean_from_monthly(monthly["date"], monthly["v"])


def ltir_canada():
    df = _boc_valet("CDN.AVG.OVER.10.AVG", FETCH_START, FETCH_END)  # >10y gov bonds (daily)
    return df.assign(year=df["date"].dt.year).groupby("year")["v"].mean()


def sr_japan():
    url = ("https://www.stat-search.boj.or.jp/api/v1/getDataCode?format=json&lang=en"
           f"&db=FM02&code=STRACLUCON&startDate={FETCH_START}01&endDate={FETCH_END}12")
    payload = json.loads(_get(url, {"Accept": "application/json"}))
    v = payload["RESULTSET"][0]["VALUES"]
    months = pd.to_datetime(v["SURVEY_DATES"], format="%Y%m")
    return _annual_mean_from_monthly(months, v["VALUES"])


def sr_belgium():
    import requests
    r = requests.get("https://nsidisseminate-stat.nbb.be/rest/data/BE2,DF_IRTRCERT,1.0/M.3M",
                     params={"startPeriod": f"{FETCH_START}-01", "format": "csv"},
                     timeout=120, verify=False, headers={"User-Agent": UA})
    r.raise_for_status()
    raw = pd.read_csv(io.StringIO(r.text))
    months = pd.to_datetime(raw["TIME_PERIOD"], format="%Y-%m")
    return _annual_mean_from_monthly(months, raw["OBS_VALUE"])


def sr_norway():
    raw = _get("https://data.norges-bank.no/api/data/SHORT_RATES/A.NOWA.ON.R?format=csv&locale=en")
    df = pd.read_csv(io.StringIO(raw.decode("utf-8", "replace")), sep=";")
    df = df[["TIME_PERIOD", "OBS_VALUE"]].copy()
    df["year"] = pd.to_numeric(df["TIME_PERIOD"], errors="coerce")
    df["v"] = pd.to_numeric(df["OBS_VALUE"], errors="coerce")
    df = df.dropna(subset=["year", "v"])
    return pd.Series(df["v"].values, index=df["year"].astype(int).values)


def sr_switzerland():
    """SNB zimoma call money (Tomorrow next); December value each year (as in notebook)."""
    url = f"https://data.snb.ch/api/cube/zimoma/data/json/en?{urllib.parse.urlencode({'fromDate': f'{FETCH_START}-01'})}"
    payload = json.loads(_get(url, {"Accept": "application/json"}))
    target = "Switzerland - CHF - Call money rate (Tomorrow next) - 1 day"
    series = next(ts for ts in payload["timeseries"] if ts["header"][0]["dimItem"] == target)
    d = pd.DataFrame({"month": pd.to_datetime([x["date"] for x in series["values"]], format="%Y-%m"),
                      "v": pd.to_numeric([x["value"] for x in series["values"]], errors="coerce")}).dropna()
    dec = d[d["month"].dt.month == 12]
    return pd.Series(dec["v"].values, index=dec["month"].dt.year.values)


def sr_australia():
    raw = pd.read_csv(RBA_CSV, usecols=[0, 1], names=["date", "v"], skiprows=11)
    months = pd.to_datetime(raw["date"], format="%d/%m/%Y", errors="coerce")
    return _annual_mean_from_monthly(months, raw["v"])


def rate_fred_graph(series_id):
    raw = _get(f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}").decode()
    df = pd.read_csv(io.StringIO(raw))
    df.columns = ["date", "v"]
    df["date"] = pd.to_datetime(df["date"])
    df["v"] = pd.to_numeric(df["v"], errors="coerce")
    df = df.dropna()
    return df.assign(year=df["date"].dt.year).groupby("year")["v"].mean()


def ltir_uk_boe():
    """Bank of England IUAALNPY 20-yr par yield via Playwright (may be unavailable)."""
    import asyncio
    import threading
    from playwright.async_api import async_playwright

    url = ("https://www.bankofengland.co.uk/boeapps/database/fromshowcolumns.asp"
           "?Travel=NIxAZxSUx&FromSeries=1&ToSeries=50&DAT=RNG&FD=1&FM=Jan&FY=2009"
           "&TD=1&TM=Jan&TY=2040&FNY=Y&CSVF=TT&html.x=66&html.y=26"
           "&SeriesCodes=IUAALNPY&UsingCodes=Y&Filter=N&title=IUAALNPY&VPD=Y")

    async def _fetch():
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])
            ctx = await browser.new_context(user_agent=UA, locale="en-GB",
                                            viewport={"width": 1280, "height": 800})
            await ctx.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
            page = await ctx.new_page()
            last = None
            for _ in range(3):
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=90000)
                    try:
                        await page.click("button[data-cookie='analytics']", timeout=4000)
                    except Exception:  # noqa
                        pass
                    await page.wait_for_selector("table tbody tr td", timeout=30000)
                    async with page.expect_download(timeout=30000) as dl:
                        await page.click(".dt-button:has-text('CSV')")
                    path = await (await dl.value).path()
                    with open(path, "rb") as f:
                        data = f.read()
                    await browser.close()
                    return data
                except Exception as exc:  # noqa
                    last = exc
                    await page.wait_for_timeout(3000)
            await browser.close()
            raise last

    box = {}

    def _run():
        # WindowsProactorEventLoopPolicy only exists on Windows (it is required
        # there for Playwright's subprocess transport). On Linux/macOS the
        # attribute does not exist, so calling it unconditionally raised
        # AttributeError in this worker thread and killed the BoE/UK long-rate
        # fetch. Guard by platform; the default loop policy is correct elsewhere.
        if sys.platform.startswith("win"):
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        box["csv"] = loop.run_until_complete(_fetch())
        loop.close()

    # daemon=True + a bounded join so a headless Playwright launch that never
    # completes (common on Linux compute nodes: Chromium never spawns from this
    # non-main-thread event loop) degrades to a graceful [FAIL] fallback instead
    # of hanging the entire pull forever. See PIPELINE_BUGS.md BUG 2.
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    t.join(timeout=150)
    if t.is_alive() or "csv" not in box:
        raise RuntimeError("BoE UK long-rate fetch did not complete within 150s "
                           "(headless Playwright hang); leaving ltir_uk unchanged")
    raw = pd.read_csv(io.BytesIO(box["csv"]))
    raw.columns = ["date", "lr"]
    raw["date"] = pd.to_datetime(raw["date"], format="%d %b %y", errors="coerce")
    raw["lr"] = pd.to_numeric(raw["lr"], errors="coerce")
    raw = raw.dropna()
    return raw.assign(year=raw["date"].dt.year).groupby("year")["lr"].mean()


def worldbank_rconpc_index(iso3):
    """World Bank NE.CON.PRVT.PC.KD rebased to RCONPC_BASE_YEAR (2010) = 100."""
    import time

    import requests
    url = (f"https://api.worldbank.org/v2/country/{iso3}/indicator/NE.CON.PRVT.PC.KD"
           f"?format=json&date={RCONPC_BASE_YEAR}:{FETCH_END}&per_page=10000")
    last = None
    for attempt in range(5):
        try:
            r = requests.get(url, timeout=60, verify=False,
                             headers={"User-Agent": UA, "Accept": "application/json"})
            r.raise_for_status()
            payload = r.json()
            if not isinstance(payload, list) or len(payload) < 2 or payload[1] is None:
                raise ValueError("empty World Bank payload")
            data = pd.json_normalize(payload[1])
            d = pd.DataFrame({"year": pd.to_numeric(data["date"], errors="coerce"),
                              "v": pd.to_numeric(data["value"], errors="coerce")}).dropna()
            s = pd.Series(d["v"].values, index=d["year"].astype(int).values).sort_index()
            return _rebase_index(s, RCONPC_BASE_YEAR)
        except Exception as exc:  # noqa
            last = exc
            time.sleep(2 * (attempt + 1))
    raise last


# --------------------------------------------------------------------------- #
# Column dispatch
# --------------------------------------------------------------------------- #
# CPI: country-suffix -> callable returning a level/index Series
CPI_SRC = {
    "usa": cpi_bls_usa, "uk": cpi_ons_uk, "fr": cpi_imf_ifs_fr,
    "ie": cpi_cso_ie, "se": cpi_scb_se, "no": cpi_norway,
}
for _c in ["deu", "ca", "it", "jp", "au", "be", "fi", "nl", "ch", "es", "pt", "dk"]:
    CPI_SRC[_c] = (lambda iso: (lambda: cpi_weo(iso)))(ISO3[_c])

# Short rate: country-suffix -> callable returning annual-rate Series
SR_SRC = {
    "usa": sr_measuringworth_usa,
    "deu": sr_germany,
    "fr": sr_france,
    "ca": sr_canada,
    "it": lambda: imf_mfs("ITA", "GSTBILY_RT_PT_A_PT"),
    "jp": sr_japan,
    "au": sr_australia,
    "be": sr_belgium,
    "fi": lambda: oecd_short("FIN"),
    "ie": lambda: oecd_short("IRL"),
    "nl": lambda: oecd_short("NLD"),
    "no": sr_norway,
    "ch": sr_switzerland,
    "se": lambda: oecd_short("SWE"),
    "es": lambda: oecd_short("ESP"),
    "pt": lambda: oecd_short("PRT"),
    "dk": lambda: imf_mm_then_oecd("DNK"),
    "uk": sr_measuringworth_uk,
}

# Long rate: country-suffix -> callable returning annual-rate Series
LR_SRC = {
    "usa": lambda: oecd_finmark("USA", "IRLT"),
    "deu": lambda: oecd_finmark("DEU", "IRLT"),
    "uk": ltir_uk_boe,
    "fr": lambda: oecd_finmark("FRA", "IRLT"),
    "ca": ltir_canada,
    "it": lambda: imf_bonds_then_oecd("ITA"),
    "jp": lambda: oecd_finmark("JPN", "IRLT"),
    "au": lambda: oecd_finmark("AUS", "IRLT"),
    "be": lambda: oecd_finmark("BEL", "IRLT"),
    "fi": lambda: oecd_finmark("FIN", "IRLT"),
    "ie": lambda: oecd_finmark("IRL", "IRLT"),
    "nl": lambda: oecd_finmark("NLD", "IRLT"),
    "no": lambda: imf_bonds_then_oecd("NOR"),
    "ch": lambda: imf_bonds_then_oecd("CHE"),
    "se": lambda: oecd_finmark("SWE", "IRLT"),
    "es": lambda: oecd_finmark("ESP", "IRLT"),
    "pt": lambda: oecd_finmark("PRT", "IRLT"),
    "dk": lambda: imf_bonds_then_oecd("DNK"),
}

COUNTRY_ORDER = ["usa", "deu", "uk", "fr", "ca", "it", "jp", "au", "be", "fi",
                 "ie", "nl", "no", "ch", "se", "es", "pt", "dk"]


# --------------------------------------------------------------------------- #
# Build the fetched values, keyed by workbook column name
# --------------------------------------------------------------------------- #
def _fetch(label, fn):
    """Run a fetch, log a coverage summary, and return the RAW year-indexed
    Series (or None on failure / no usable 2017-2025 data)."""
    try:
        s = fn().sort_index()
        window = s[(s.index >= FETCH_START) & (s.index <= FETCH_END)].dropna()
        if len(window) == 0:
            print(f"  [warn ] {label:12s}: no 2017-2025 data -> keeping existing")
            return None
        print(f"  [ ok  ] {label:12s}: {int(window.index.min())}-{int(window.index.max())} "
              f"({len(window)} yrs)")
        return s
    except Exception as exc:  # noqa
        print(f"  [FAIL ] {label:12s}: {type(exc).__name__}: {exc} -> keeping existing")
        return None


def _place_index_with_reanchor(existing, col, level_series):
    """Place source index levels for 2017..END.

    2017..CHAIN_REANCHOR_YEAR: source levels directly.
    If source vs existing at CHAIN_REANCHOR_YEAR differs by > CHAIN_REANCHOR_PCT,
    pin that year to the existing workbook and chain 2021+ from it using source
    year-over-year ratios.  Otherwise place source levels through FETCH_END.
    """
    full = level_series.dropna().sort_index()
    if full.empty:
        return None

    vals = {}
    for y in range(FETCH_START, CHAIN_REANCHOR_YEAR + 1):
        if y in full.index:
            vals[y] = float(full[y])

    old_anchor = (existing.at[CHAIN_REANCHOR_YEAR, col]
                  if CHAIN_REANCHOR_YEAR in existing.index else None)
    new_anchor = vals.get(CHAIN_REANCHOR_YEAR)
    reanchor = False

    if (new_anchor is not None and old_anchor is not None and pd.notna(old_anchor)
            and pd.notna(new_anchor) and abs(old_anchor) > 1e-9):
        pct = abs((new_anchor - old_anchor) / old_anchor) * 100
        if pct > CHAIN_REANCHOR_PCT:
            reanchor = True
            vals[CHAIN_REANCHOR_YEAR] = float(old_anchor)
            level, prev_year = float(old_anchor), CHAIN_REANCHOR_YEAR
            for y in range(CHAIN_REANCHOR_YEAR + 1, FETCH_END + 1):
                if y in full.index and prev_year in full.index:
                    level = level * (full[y] / full[prev_year])
                    vals[y] = level
                    prev_year = y
            print(f"         {col}: {CHAIN_REANCHOR_YEAR} diff {pct:.2f}% > "
                  f"{CHAIN_REANCHOR_PCT}% -> chain 2021+ from existing {CHAIN_REANCHOR_YEAR}")

    if not reanchor:
        for y in range(CHAIN_REANCHOR_YEAR + 1, FETCH_END + 1):
            if y in full.index:
                vals[y] = float(full[y])

    return vals or None


def main():
    print(f"Reading source workbook: {SRC_XLSX}")
    existing = pd.read_excel(SRC_XLSX, sheet_name="Data")
    existing = existing[pd.to_numeric(existing["year"], errors="coerce").notna()].copy()
    existing["year"] = existing["year"].astype(int)
    existing = existing.set_index("year")

    # collected: column-name -> {year: value} to write for 2017..2025
    write = {}

    def _place(col, s):
        """Place fetched values directly for 2017..END (blank where absent)."""
        w = s[(s.index >= FETCH_START) & (s.index <= FETCH_END)].dropna()
        write[col] = {int(y): float(v) for y, v in w.items()}

    print(f"\n=== CPI (source {CPI_BASE_YEAR}=100; re-chain at {CHAIN_REANCHOR_YEAR} if >{CHAIN_REANCHOR_PCT}%) ===")
    for c in COUNTRY_ORDER:
        col = f"cpi_{c}"
        s = _fetch(col, CPI_SRC[c])
        if s is None:
            continue
        vals = _place_index_with_reanchor(existing, col, s)
        if vals:
            write[col] = vals

    print("\n=== Short-term rates (placed directly) ===")
    for c in COUNTRY_ORDER:
        if c not in SR_SRC:
            print(f"  [keep ] stir_{c}: no source -> existing values retained")
            continue
        s = _fetch(f"stir_{c}", SR_SRC[c])
        if s is not None:
            _place(f"stir_{c}", s)

    print("\n=== Long-term rates (placed directly) ===")
    for c in COUNTRY_ORDER:
        if c not in LR_SRC:
            print(f"  [keep ] ltir_{c}: no source -> existing values retained")
            continue
        s = _fetch(f"ltir_{c}", LR_SRC[c])
        if s is not None:
            _place(f"ltir_{c}", s)

    print("\n=== baa_usa (placed directly) ===")
    s = _fetch("baa_usa", lambda: rate_fred_graph("BAA"))
    if s is not None:
        _place("baa_usa", s)

    print(f"\n=== Real consumption per capita (World Bank {RCONPC_BASE_YEAR}=100; "
          f"re-chain at {CHAIN_REANCHOR_YEAR} if >{CHAIN_REANCHOR_PCT}%) ===")
    for c in COUNTRY_ORDER:
        col = f"rconpc_{c}"
        s = _fetch(col, lambda cc=c: worldbank_rconpc_index(ISO3[cc]))
        if s is None:
            continue
        vals = _place_index_with_reanchor(existing, col, s)
        if vals:
            write[col] = vals

    # --------------------------------------------------------------------- #
    # Write into a copy of the original workbook (preserves Sheet1 + Spec)
    # --------------------------------------------------------------------- #
    import openpyxl
    print(f"\nWriting output workbook: {OUT_XLSX}")
    wb = openpyxl.load_workbook(SRC_XLSX)
    ws = wb["Data"]

    header = {str(ws.cell(row=1, column=col).value): col
              for col in range(1, ws.max_column + 1)}
    year_row = {}
    for row in range(2, ws.max_row + 1):
        yv = ws.cell(row=row, column=1).value
        try:
            year_row[int(yv)] = row
        except (TypeError, ValueError):
            continue

    # ensure a row exists for every fetch year (append new years like 2025)
    next_row = max(year_row.values()) + 1
    for y in range(FETCH_START, FETCH_END + 1):
        if y not in year_row:
            ws.cell(row=next_row, column=1, value=y)
            year_row[y] = next_row
            next_row += 1

    n_cells = 0
    for col_name, vals in write.items():
        if col_name not in header:
            print(f"  (skip: {col_name} not a workbook column)")
            continue
        col = header[col_name]
        for y in range(FETCH_START, FETCH_END + 1):
            cell = ws.cell(row=year_row[y], column=col)
            cell.value = float(vals[y]) if y in vals else None
            n_cells += 1

    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")

    # Strip rconpc_* columns for the shorter workbook (Sheet1 + Spec unchanged).
    wb_norc = openpyxl.load_workbook(OUT_XLSX)
    ws_norc = wb_norc["Data"]
    drop_cols = [col for col in range(1, ws_norc.max_column + 1)
                 if str(ws_norc.cell(row=1, column=col).value or "").startswith("rconpc_")]
    for col in sorted(drop_cols, reverse=True):
        ws_norc.delete_cols(col)
    wb_norc.save(OUT_XLSX_NO_RCONPC)
    print(f"Saved: {OUT_XLSX_NO_RCONPC} ({len(drop_cols)} rconpc columns removed)")

    print(f"Done. Overwrote {n_cells} cells across {len(write)} columns "
          f"for {FETCH_START}-{FETCH_END}.")
    print("Columns refreshed:", ", ".join(sorted(write)))


if __name__ == "__main__":
    sys.exit(main())
